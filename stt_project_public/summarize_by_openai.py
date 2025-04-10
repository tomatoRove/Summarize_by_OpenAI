import glob
import openpyxl
import re
import openai
from openai import OpenAI
import os


def main():
    # Specify the Excel to write in
    target_excel = r"<target_excel's file path>"
    try:
        workbook = openpyxl.load_workbook(target_excel)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        print("The specified Excel file cannot be found. A new file has been created.")    
    
    number_of_row = int(input("Enter the first number of the folder.")) + 1    

    # *_original.txt or *_output.txt
    for stt_file_path in glob.glob(r'*_output.txt'):
        try:
            with open(stt_file_path, 'r', encoding='utf-8') as file:
                target_text = file.read() 
        except FileNotFoundError:
            print(f"'{stt_file_path}' not found.")
        except Exception as e:
            print(f"An error has occurred: {e}")

        summarize_text = summarize_text_by_AI(target_text)
        
        
        if "original" in stt_file_path:     # Enter summary results of original file on a worksheet
            head_number_file = re.search(r"(\d+)_original.txt", stt_file_path)
            number_of_row = int(head_number_file.group(1)) + 1
            print(f"Start to process No.0{number_of_row}.")
            result_to_excel_originalVER(workbook, summarize_text, number_of_row)
            print(f"Complete to process No.0{number_of_row}.")
        else:                               # Enter summary results of STT outputs on worksheets for each sercices
            result_to_excel(workbook, stt_file_path, summarize_text, number_of_row)
        
        workbook.save(target_excel)


def summarize_text_by_AI(target_text):
    openai.api_key = "<Your API key>"
    client = OpenAI()

    # AI settings (Caution: Some models require the use of different grammars.)
    response = client.chat.completions.create(
        model = "gpt-4o", 
        messages = [
            {"role": "system", "content": f"次のテキストを3行に要約してください。だ・である調かつ数字付きリスト形式で出力してください。出力内容に創作を含めることを禁じます。"},      # An order to AI
            {"role": "user", "content": f"{target_text}"},      # Target content
        ],
    )

    # Summarize result
    summary = response.choices[0].message.content
    print(f"summary:")
    print(f"{summary}")
    
    return summary

def get_column_number(search_word, work_sheet):
    for i in range(1, work_sheet.max_column + 1):
        header_cell = work_sheet.cell(row=1, column=i).value
        if header_cell == search_word:
            return i
    return print(f"{search_word}'s column number not found.")

def result_to_excel_originalVER(workbook, summarize_text, number_of_row):
    # List of STT services
    list_sheet_name = ["speechmatics_standard", "speechmatics_enhanced"]

    for sheet_name in list_sheet_name:
        work_sheet = workbook[sheet_name]

        print(f"Start writing to {sheet_name}")

        number_of_column = get_column_number("AI_Summarization_Original", work_sheet) 
        
        work_sheet.cell(row=number_of_row, column=number_of_column).value = summarize_text

        print(f"Complete writing to {sheet_name}")


def result_to_excel(workbook, stt_file_path, summarize_result, number_of_row):
    # Get STT service's name to specify the worksheet.
    sheet_name = re.search(r'(\w+)_output.txt', stt_file_path)
    if sheet_name:
        sheet_name = sheet_name.group(1)
    else:
        print("Cannot get sheet name.")
        return

    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist.")
        return
    else:
        print(f"Start writing to {sheet_name}")

    work_sheet = workbook[sheet_name]

    # Write the summarize results to Excel
    number_of_column = get_column_number("AI_Summarization_STT", work_sheet)
    work_sheet.cell(row=number_of_row, column=number_of_column).value = summarize_result
    
    print(f"Complete writing to {sheet_name}.")
    return

    

if __name__ == '__main__':
    main()
